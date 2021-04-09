import telebot
import openpyxl

wb = openpyxl.load_workbook('sotr/2020.xlsx')

bot = telebot.TeleBot('1530619842:AAH20gD_rs72BX9HsXlef2avhbxMPXUiROk')
keyboard1 = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
keyboard1.row('новости', 'объявления')
keyboard1.row('список сотрудников', 'документы')
keyboard1.row('вакансии', 'контакты')

sheet = wb.active
rows = sheet.max_row
cols = sheet.max_column

def printsotr(txt):
    string = ''
    count = 0
    if len(txt.text) > 2:
        bot.reply_to(txt, 'Две буквы, а не больше!')
        return
    firstletter = (txt.text[0].upper())
    secondletter = (txt.text[1].lower())
    twofirstletter = (firstletter + secondletter)
    for i in range(1, rows + 1):
        cell = sheet.cell(row=i, column=1)
        column = sheet.cell(row=i, column=2)
        if (twofirstletter) == (cell.value[:2]):
            count += 1
            bot.reply_to(txt, cell.value + ' : ' + column.value)
    if count == 0:
        bot.reply_to(txt, 'Нет такой фамилии')


@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, 'Привет, ты написал мне /start', reply_markup=keyboard1)

@bot.message_handler(content_types=['text'])
def send_text(message):
    if message.text.lower() == 'привет':
        bot.send_message(message.chat.id, 'Привет')
    elif message.text.lower() == 'пока':
        bot.send_message(message.chat.id, 'Прощай')
    elif message.text.lower() == 'новости':
        bot.send_message(message.chat.id, 'У Казанцева отпуск с 5 апреля')
    elif message.text.lower() == 'список сотрудников':
        msg = bot.send_message(message.chat.id, 'Введите 2 первые буквы фамилии:')
        bot.register_next_step_handler(msg, printsotr)
    elif message.text.lower() == 'документы':
        bot.send_message(message.chat.id, 'нет доступа')
    elif message.text.lower() == 'объявления':
        bot.send_message(message.chat.id, 'в разработке')
    elif message.text.lower() == 'пошел':
        bot.send_message(message.chat.id, 'CAADAgADZgkAAnlc4gmfCor5YbYYRAI')
    elif message.text.lower() == 'вакансии':
        bot.send_message(message.chat.id, 'Электросварщик на автоматических и полуавтоматических машинах')
        bot.send_message(message.chat.id, 'Контролер ОТК с функциями кладовщика')
        bot.send_message(message.chat.id, 'Инженер - технолог')
        bot.send_message(message.chat.id, 'Термист')
        bot.send_message(message.chat.id, 'Наладчик с танков с ЧПУ')
        bot.send_message(message.chat.id, 'Мастер цеха металлообработки')
    elif message.text.lower() == 'контакты':
        bot.send_message(message.chat.id, '+7 (343) 221-53-54 – офис ')
        bot.send_message(message.chat.id, '+7 (343) 221-53-55 – бухгалтерия ')
        bot.send_message(message.chat.id, 'Адрес 620085, Екатеринбург, ул. Монтерская, 3, цех 11 (территория ОАО "Уральский завод резиновых технических изделий"')

bot.polling()